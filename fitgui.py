from PyQt5.QtWidgets import *
from matplotlib.pyplot import show
from typing import List, Union
from load_data_v2 import LoadData
from fit import Fit
import importlib.util
import sys
from pathlib import Path
from typing import TYPE_CHECKING
import json
from os.path import exists
from PyQt5.QtGui import QFont

if TYPE_CHECKING:
    import types

help_data = '* Data file must be an Excel file or a CSV file.\n'
help_model = '* Every fitting function MUST be written in a different python script.\n' \
             '\n* The name of the script is irrelevant to the operation of the code, different fitting function that are written should be identifyable by the script name.\n' \
             '\n* The name of the function MUST be PRECISELY "fit_function".\n' \
             '\n* The ODR algorithm and the Least Squares algorithm require different definition of the fitting function:\n' \
             '\n1. The ODR algorithm requires that the fitting function takes as the first argument a parameters VECTOR (a vector containing the parameters which the algorithm then finds the best fitting parameters) and X as the second argument:\n' \
             '\n\tdef fit_function(a, x):\n' \
             '\t\treturn a[0] * x + a[1]\n' \
             '\n2. The LeastSquares algorithm requires that the fitting function takes X as the first argument and the fitting parameters will be all the rest arguments:\n' \
             '\n\tdef fit_function(x, a, b):\n' \
             '\t\treturn a * x + b\n' \
             '\n* Please refer to the attached example scripts: "example_linear_odr.py", "example_linear_least_squares.py"\n'
help_method = 'The ODR algorithm takes into account the errors in X,\n' \
              'where as the Least Squares one does not.\n' \
              'If the errors in the X axis are not important\n' \
              'the Least Squares algorithm is recommended'
help_cols = '* Does your data sheet have the first row as names for each column?\n' \
            '\tIf so check the "Headers" checkbox.\n' \
            '\n* Giving names to each column is highly recommended and is considered a good practice, hence it is checked by default.\n' \
            '\n* The columns are 0 indexed, i.e the first column is indexed as 0.\n' \
            '\n* Different columns CAN NOT have the same index.\n' \
            '\n* For the Least Squares algorithm the dX column is not used.'
help_del = '* Do you want to remove points from your data set?\n' \
           '\tIf so write the elements of the array which you would like to remove and check the checkbox.\n' \
           '\n* Removing measurments is frowned uppon in the scientific community, you should be very careful and have a very good reason to do so\n' \
           '\t* YOU HAVE BEEN WARNED!\n' \
           '\n* The elements are 0 indexed, i.e the first element is indexed as 0.\n' \
           '\n* The elements must be referred by an INTEGER and not by a float.\n' \
           '\n* The input must match the Pythonic listing standards:\n' \
           '\tThe elements must be seperated by a comma and exactly one space ", ":\n' \
           '\tFor example: if you want to remove the first 3 points you would write:\n' \
           '\t\t"0, 1, 2"'
help_initial = '* The number of the initial parameters that you provide\n' \
               'MUST match the number of the parameters which are defined in the "fit_function"\n' \
               'and in the order in which they are defined.\n' \
               '\nFor example, if the function is defined as such (for the ODR algorithm):\n' \
               '\n\tdef fit_function(a, x):' \
               '\t\treturn a[0] * x + a[1]\n' \
               '\nOr as such (for the Least Squares algorithm):\n' \
               '\n\tdef fit_function(x, a, b):\n' \
               '\t\treturn a * x + b\n' \
               '\nAnd the initial parameters "1, 2":\n' \
               '\t* a[0]<=>a<=>1\n' \
               '\t* a[1]<=>b<=>2\n' \
               '\n* The input must match the Pythonic listing standards:\n' \
               '\tThe elements must be seperated by a comma and exactly one space ", ":\n' \
               '\tFor example: if you want to give the program "1" as the first parameter\n' \
               '\tand "2" as the second, you would write:\n' \
               '\t\t"1, 2"'
help_labels = '* The program uses matplotlib to plot, and so accepts (only) Latex syntax\n' \
              '\n * Please refer to the following site for Latex symbols:\n' \
              '\nhttps://oeis.org/wiki/List_of_LaTeX_mathematical_symbols\n' \
              '\n* One important note:\n' \
              '\tTo write "Space" (i.e " ") in Latex you would write a Backward Slash and one Space "\ ".\n' \
              '\tFor example, to display "Hellow World" inside the plot, you would enter in Latex syntax: "Hello\ World".'


def import_source_file(fname: Union[str, Path], modname: str) -> "types.ModuleType":
    """
     Import a Python source file and return the loaded module.

     Args:
         fname: The full path to the source file.  It may container characters like `.`
             or `-`.
         modname: The name for the loaded module.  It may contain `.` and even characters
             that would normally not be allowed (e.g., `-`).
     Return:
         The imported module

     Raises:
         ImportError: If the file cannot be imported (e.g, if it's not a `.py` file or if
             it does not exist).
         Exception: Any exception that is raised while executing the module (e.g.,
             :exc:`SyntaxError).  These are errors made by the author of the module!
     """
    # https://docs.python.org/3/library/importlib.html#importing-a-source-file-directly
    spec = importlib.util.spec_from_file_location(modname, fname)
    if spec is None:
        raise ImportError(f"Could not load spec for module '{modname}' at: {fname}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[modname] = module
    try:
        spec.loader.exec_module(module)
    except FileNotFoundError as e:
        raise ImportError(f"{e.strerror}: {fname}") from e
    return module


class FitGUI(QMainWindow):
    def setup_ui(self) -> None:

        font = QFont()
        font.setPointSize(10)

        self.centralwidget = QWidget(self)
        self.centralwidget.setFont(font)

        grid = QGridLayout()
        grid.setSpacing(10)

        # Adding all the widgets to the central widget

        # 1'st Row
        self.label_loaddata = QLabel(self.centralwidget)
        self.label_loaddata.setText('Load Data File:')
        grid.addWidget(self.label_loaddata, 0, 0)

        self.toolButton_help_data_file = QToolButton(self.centralwidget)
        self.toolButton_help_data_file.setText('?')
        grid.addWidget(self.toolButton_help_data_file, 0, 1)

        self.lineEdit_pathdata = QLineEdit(self.centralwidget)
        self.lineEdit_pathdata.setReadOnly(True)
        grid.addWidget(self.lineEdit_pathdata, 0, 2, 1, 6)  # row 0, col 1, 1 rowspan, 2 colspan

        self.pushButton_browsedata = QPushButton(self.centralwidget)
        self.pushButton_browsedata.setText('Browse')
        grid.addWidget(self.pushButton_browsedata, 0, 8, 1, 2)

        # 2'nd Row
        self.label_loadmodel = QLabel(self.centralwidget)
        self.label_loadmodel.setText('Load Model File:')
        grid.addWidget(self.label_loadmodel, 1, 0)

        self.toolButton_help_model_file = QToolButton(self.centralwidget)
        self.toolButton_help_model_file.setText('?')
        grid.addWidget(self.toolButton_help_model_file, 1, 1)

        self.lineEdit_pathmodel = QLineEdit(self.centralwidget)
        self.lineEdit_pathmodel.setReadOnly(True)
        grid.addWidget(self.lineEdit_pathmodel, 1, 2, 1, 6)

        self.pushButton_browsemodel = QPushButton(self.centralwidget)
        self.pushButton_browsemodel.setText('Browse')
        grid.addWidget(self.pushButton_browsemodel, 1, 8, 1, 2)

        # 3'rd Row only appears if the file is xlsx and have sheets
        self.label_sheets = QLabel(self.centralwidget)
        self.label_sheets.setText('Sheet:')
        grid.addWidget(self.label_sheets, 2, 0)
        self.label_sheets.hide()
        # Will only append items ones the file is loaded
        self.comboBox_sheets = QComboBox(self.centralwidget)
        grid.addWidget(self.comboBox_sheets, 2, 2, 1, 2)
        self.comboBox_sheets.hide()

        # 4'th row
        self.label_method = QLabel(self.centralwidget)
        self.label_method.setText('Method:')
        grid.addWidget(self.label_method, 3, 0)

        self.toolButton_help_method = QToolButton(self.centralwidget)
        self.toolButton_help_method.setText('?')
        grid.addWidget(self.toolButton_help_method, 3, 1)

        self.comboBox_method = QComboBox(self.centralwidget)
        self.comboBox_method.addItems(['ODR', 'Least Squares'])
        grid.addWidget(self.comboBox_method, 3, 2, 1, 2)

        # 5'th row
        self.checkBox_headers = QCheckBox(self.centralwidget)
        self.checkBox_headers.setText('Headers')
        self.checkBox_headers.setChecked(True)
        grid.addWidget(self.checkBox_headers, 4, 0)

        self.toolButton_help_cols = QToolButton(self.centralwidget)
        self.toolButton_help_cols.setText('?')
        grid.addWidget(self.toolButton_help_cols, 4, 1)

        self.label_xcol = QLabel(self.centralwidget)
        self.label_xcol.setText('X Column:')
        grid.addWidget(self.label_xcol, 4, 2)

        self.spinBox_xcol = QSpinBox(self.centralwidget)
        grid.addWidget(self.spinBox_xcol, 4, 3)

        self.label_dxcol = QLabel(self.centralwidget)
        self.label_dxcol.setText('dX Column:')
        grid.addWidget(self.label_dxcol, 4, 4)

        self.spinBox_dxcol = QSpinBox(self.centralwidget)
        grid.addWidget(self.spinBox_dxcol, 4, 5)

        self.label_ycol = QLabel(self.centralwidget)
        self.label_ycol.setText('Y Column:')
        grid.addWidget(self.label_ycol, 4, 6)

        self.spinBox_ycol = QSpinBox(self.centralwidget)
        grid.addWidget(self.spinBox_ycol, 4, 7)

        self.label_dycol = QLabel(self.centralwidget)
        self.label_dycol.setText('dY Column:')
        grid.addWidget(self.label_dycol, 4, 8)

        self.spinBox_dycol = QSpinBox(self.centralwidget)
        grid.addWidget(self.spinBox_dycol, 4, 9)

        # 6'th row
        self.checkBox_delpoints = QCheckBox(self.centralwidget)
        self.checkBox_delpoints.setText('Delete Points?')
        grid.addWidget(self.checkBox_delpoints, 5, 0)

        self.toolButton_help_points = QToolButton(self.centralwidget)
        self.toolButton_help_points.setText('?')
        grid.addWidget(self.toolButton_help_points, 5, 1)

        self.lineEdit_listpoints = QLineEdit(self.centralwidget)
        grid.addWidget(self.lineEdit_listpoints, 5, 2, 1, 3)

        # 7'th row
        self.label_params = QLabel(self.centralwidget)
        self.label_params.setText('Initial Parameters:')
        grid.addWidget(self.label_params, 6, 0)

        self.toolButton_help_params = QToolButton(self.centralwidget)
        self.toolButton_help_params.setText('?')
        grid.addWidget(self.toolButton_help_params, 6, 1)

        self.lineEdit_params = QLineEdit(self.centralwidget)
        grid.addWidget(self.lineEdit_params, 6, 2, 1, 3)

        # 8'th row
        self.label_labels = QLabel(self.centralwidget)
        self.label_labels.setText('Labels:')
        font0 = QFont()
        font0.setPointSize(14)
        self.label_labels.setFont(font0)
        grid.addWidget(self.label_labels, 7, 0)

        self.toolButton_help_labels = QToolButton(self.centralwidget)
        self.toolButton_help_labels.setText('?')
        grid.addWidget(self.toolButton_help_labels, 7, 1)

        # 9'th row
        self.label_fittitle = QLabel(self.centralwidget)
        self.label_fittitle.setText('Fit Title:')
        grid.addWidget(self.label_fittitle, 8, 0)

        self.lineEdit_fittitle = QLineEdit(self.centralwidget)
        grid.addWidget(self.lineEdit_fittitle, 8, 2, 1, 3)

        self.checkBox_fit = QCheckBox(self.centralwidget)
        self.checkBox_fit.setChecked(True)
        self.checkBox_fit.setText('Plot fit')
        grid.addWidget(self.checkBox_fit, 8, 6, 1, 3)

        # 10'th Row
        self.label_fitxlabel = QLabel(self.centralwidget)
        self.label_fitxlabel.setText('Fit X Label:')
        grid.addWidget(self.label_fitxlabel, 9, 0)

        self.lineEdit_fitxlabel = QLineEdit(self.centralwidget)
        grid.addWidget(self.lineEdit_fitxlabel, 9, 2, 1, 3)

        self.checkBox_residuals = QCheckBox(self.centralwidget)
        self.checkBox_residuals.setChecked(True)
        self.checkBox_residuals.setText('Plot Residuals')
        grid.addWidget(self.checkBox_residuals, 9, 6, 1, 3)

        # 11'th Row
        self.label_fitylabel = QLabel(self.centralwidget)
        self.label_fitylabel.setText('Fit Y Label:')
        grid.addWidget(self.label_fitylabel, 10, 0)

        self.lineEdit_fitylabel = QLineEdit(self.centralwidget)
        grid.addWidget(self.lineEdit_fitylabel, 10, 2, 1, 3)

        self.checkBox_initguess = QCheckBox(self.centralwidget)
        self.checkBox_initguess.setText('Plot Initial Guess')
        grid.addWidget(self.checkBox_initguess, 10, 6, 1, 3)

        # 12'th Row
        self.label_resylabel = QLabel(self.centralwidget)
        self.label_resylabel.setText('Residuals Y Label:')
        grid.addWidget(self.label_resylabel, 11, 0)

        self.lineEdit_resylabel = QLineEdit(self.centralwidget)
        grid.addWidget(self.lineEdit_resylabel, 11, 2, 1, 3)

        self.label_hintresylabel = QLabel(self.centralwidget)
        self.label_hintresylabel.setText('(y - y(x))')
        grid.addWidget(self.label_hintresylabel, 11, 5, 1, 2)

        # 13'th row
        self.pushButton_fitresults = QPushButton(self.centralwidget)
        self.pushButton_fitresults.setText('Fit Results')
        self.pushButton_fitresults.setEnabled(True)
        self.pushButton_fitresults.setCheckable(False)
        self.pushButton_fitresults.setChecked(False)
        self.pushButton_fitresults.hide()
        grid.addWidget(self.pushButton_fitresults, 12, 0, 1, 2)  # TODO find better column for this button

        # 14'th Row
        self.pushButton_fit = QPushButton(self.centralwidget)
        self.pushButton_fit.setText('Fit!')
        grid.addWidget(self.pushButton_fit, 13, 0, 1, 10)

        # Setting central widget
        self.centralwidget.setLayout(grid)
        self.setCentralWidget(self.centralwidget)

    def add_menubar(self) -> None:
        self.menubar = QMenuBar(self)

        self.menuOpen = QMenu(self.menubar)
        self.menuOpen.setTitle('Open')

        self.menuRun = QMenu(self.menubar)
        self.menuRun.setTitle('Run')

        self.menuDefault_Path = QMenu(self.menuOpen)
        self.menuDefault_Path.setTitle('Default Path')

        self.setMenuBar(self.menubar)

        self.statusbar = QStatusBar(self)
        self.setStatusBar(self.statusbar)

        self.actionLoad_Data_File = QAction(self)
        self.actionLoad_Data_File.setText('Load Data File')
        self.actionLoad_Data_File.setShortcut('Ctrl+O')

        self.actionLoad_Model_File = QAction(self)
        self.actionLoad_Model_File.setText('Load Model File')
        self.actionLoad_Model_File.setShortcut('Ctrl+Shift+O')

        self.actionFit = QAction(self)
        self.actionFit.setText('Fit')
        self.actionFit.setShortcut('Ctrl+Return')

        self.actionSet_Default_Data_Path = QAction(self)
        self.actionSet_Default_ODR_Model_Path = QAction(self)
        self.actionSet_Default_Least_Squares_Model_Path = QAction(self)

        self.menuDefault_Path.addAction(self.actionSet_Default_Data_Path)
        self.menuDefault_Path.addAction(self.actionSet_Default_ODR_Model_Path)
        self.menuDefault_Path.addAction(self.actionSet_Default_Least_Squares_Model_Path)

        self.actionSet_Default_Data_Path.setText("Set Default Data Path")
        self.actionSet_Default_ODR_Model_Path.setText("Set Default ODR Model Path")
        self.actionSet_Default_Least_Squares_Model_Path.setText("Set Default Least Squares Model Path")

        self.menuOpen.addAction(self.actionLoad_Data_File)
        self.menuOpen.addAction(self.actionLoad_Model_File)
        self.menuOpen.addSeparator()
        self.menuOpen.addAction(self.menuDefault_Path.menuAction())

        self.menuRun.addAction(self.actionFit)

        self.menubar.addAction(self.menuRun.menuAction())
        self.menubar.addAction(self.menuOpen.menuAction())

    def add_functionality(self) -> None:

        self.checkBox_delpoints.toggled['bool'].connect(self.lineEdit_listpoints.setEnabled)

        self.actionSet_Default_Data_Path.triggered.connect(lambda: self.set_default_path('Data'))
        self.actionSet_Default_ODR_Model_Path.triggered.connect(lambda: self.set_default_path('ODR'))
        self.actionSet_Default_Least_Squares_Model_Path.triggered.connect(
            lambda: self.set_default_path('Least Squares'))

        self.pushButton_browsedata.clicked.connect(self.browsefilesdata)
        self.pushButton_browsemodel.clicked.connect(self.browsefilesmodel)

        self.actionLoad_Data_File.triggered.connect(self.browsefilesdata)
        self.actionLoad_Model_File.triggered.connect(self.browsefilesmodel)

        self.comboBox_method.currentIndexChanged.connect(self.disable_dx)

        self.pushButton_fit.clicked.connect(self.fit)
        self.actionFit.triggered.connect(self.fit)

        self.pushButton_fitresults.clicked.connect(self.results_window.show)

        # Help buttons

        self.toolButton_help_data_file.clicked.connect(lambda: self.popupmsg(help_data, 'help'))
        self.toolButton_help_model_file.clicked.connect(lambda: self.popupmsg(help_model, 'help'))
        self.toolButton_help_method.clicked.connect(lambda: self.popupmsg(help_method, 'help'))
        self.toolButton_help_cols.clicked.connect(lambda: self.popupmsg(help_cols, 'help'))
        self.toolButton_help_points.clicked.connect(lambda: self.popupmsg(help_del, 'help'))
        self.toolButton_help_params.clicked.connect(lambda: self.popupmsg(help_initial, 'help'))
        self.toolButton_help_labels.clicked.connect(lambda: self.popupmsg(help_labels, 'help'))

    def setup_results_window(self) -> None:
        self.results_window = QWidget()

        self.results_window.setWindowTitle('Fit Results')
        self.results_window.setMinimumSize(450, 800)

        self.results_layout = QVBoxLayout()

        # The rest of the buttons are inside FitGUI __init__
        self.results_textbox = QTextEdit()
        self.results_textbox.setReadOnly(True)
        self.results_layout.addWidget(self.results_textbox)

        # When 2 or more functions (slots) are connected to one button press (signal),
        # the functions (slots) are called in the order in which they were defined
        self.clear_button = QPushButton('Clear History')
        self.clear_button.clicked.connect(self.clear_history)
        self.clear_button.clicked.connect(self.results_window.hide)
        self.clear_button.clicked.connect(self.pushButton_fitresults.hide)
        self.results_layout.addWidget(self.clear_button)

        self.save_fit_results_button = QPushButton('Save')
        self.save_fit_results_button.clicked.connect(self.save_results_txt)
        self.results_layout.addWidget(self.save_fit_results_button)

        self.results_window.setLayout(self.results_layout)

    def __init__(self) -> None:
        super(FitGUI, self).__init__()

        self.setWindowTitle('FitGUI by Alon Ner-Gaon')

        self.setup_ui()

        self.add_menubar()

        self.setup_results_window()

        self.add_functionality()

        # Loading configuration
        self.config_path = 'Data/Config/config.json'

        self.default_data_path = None
        self.default_odr_path = None
        self.default_ls_path = None
        if exists(self.config_path):
            with open(self.config_path, 'r') as f:
                self.config = json.load(f)

            if 'Data' in self.config:
                self.default_data_path = self.config['Data']
            if 'ODR' in self.config:
                self.default_odr_path = self.config['ODR']
            if 'Least Squares' in self.config:
                self.default_ls_path = self.config['Least Squares']

        self.fit_number = 0

        self.show()

    def fit(self) -> None:
        try:
            self.check_empty_fields()

            self.fit_number += 1

            self.method = self.get_method()

            self.load_fit_function()

            headers = self.checkBox_headers.isChecked()

            self.xcol = self.spinBox_xcol.value()
            self.ycol = self.spinBox_ycol.value()
            self.dycol = self.spinBox_dycol.value()
            if self.method == 'odr':
                self.dxcol = self.spinBox_dxcol.value()
            else:
                self.dxcol = None  # if method == 'ls' the Fit class doesn't even access colorder: List[int] [1] (the
                # second entry in the colorder list)

            self.check_identical_cols_nums()

            delpoints = self.checkBox_delpoints.isChecked()
            points_to_remove: List[str] = self.lineEdit_listpoints.text().split(', ')
            indices_to_remove = None
            if delpoints:
                indices_to_remove = [int(p) for p in points_to_remove]

            data_file_ext = self.get_data_file_ext()

            if data_file_ext in ['.xlsx', '.xlsm']:

                data = LoadData(path=self.lineEdit_pathdata.text(),
                                indices_to_remove=indices_to_remove,
                                headers=headers,
                                delete_points=delpoints,
                                sheet_name=self.comboBox_sheets.currentText()).data
            else:
                data = LoadData(path=self.lineEdit_pathdata.text(),
                                indices_to_remove=indices_to_remove,
                                headers=headers,
                                delete_points=delpoints).data

            # if delpoints:
            #     data = LoadData(self.lineEdit_pathdata.text(), [int(p) for p in points_to_remove], headers,
            #                     delpoints).data
            # else:
            #     data = LoadData(self.lineEdit_pathdata.text(), headers=headers).data

            init_params: List[str] = self.lineEdit_params.text().split(', ')

            self.fit = Fit(
                data,
                [self.xcol, self.dxcol, self.ycol, self.dycol],
                [float(p) for p in init_params],
                self.fit_function,
                self.method
            )

            title = self.lineEdit_fittitle.text()
            xlabel = self.lineEdit_fitxlabel.text()
            ylabel = self.lineEdit_fitylabel.text()
            residuals_ylabel = self.lineEdit_resylabel.text()

            if self.checkBox_fit.isChecked():
                self.fit.plot_fit(title, xlabel, ylabel, self.fit_number)
            if self.checkBox_residuals.isChecked():
                self.fit.plot_residuals(xlabel, residuals_ylabel, self.fit_number)
            if self.checkBox_initguess.isChecked():
                self.fit.plot_initial_guess(xlabel, ylabel, self.fit_number)

            self.apply_fit_number(self.fit_number)

            self.results_textbox.append('\nFile: ' + self.get_fit_function_file_name() + '\n')

            self.results_textbox.append(self.fit.__str__())

            self.apply_partition()

            self.pushButton_fitresults.show()

            show()  # Show plotted graphs

        except IndexError as e:
            self.popupmsg("Most common error:\n"
                          "The number of the provided initial parameters do not match the number which is defined by 'fit_function'.\n"
                          "Error Type:\n" + "\n" + type(e).__name__ + "\n" +
                          "\n---------------------------------\n" +
                          "\nError Message:\n' + '\n" + str(e),
                          'Error')
        except TypeError as e:
            self.popupmsg("Most common error:\n"
                          "Method does not match the Model file.\n"
                          "Error Type:\n" + "\n" + type(e).__name__ + "\n" +
                          "\n---------------------------------\n" +
                          "\nError Message:\n' + '\n" + str(e),
                          'Error')

        except Exception as e:
            self.popupmsg('Error Type:\n' + '\n' + type(e).__name__ + '\n' +
                          '\n---------------------------------\n' +
                          '\nError Message:\n' + '\n' + str(e), 'error')

    def apply_fit_number(self, n) -> None:
        string = f'Fit Number {n}:\n'
        self.results_textbox.append(string)

    def apply_partition(self) -> None:
        self.results_textbox.append(
            '----------------------------------------------------------------------------------------\n'
            '----------------------------------------------------------------------------------------\n'
            '\n')

    def save_results_txt(self) -> None:
        path = QFileDialog.getSaveFileName(self, 'Save File', 'Data/fit_results.txt', 'TXT (*.txt)')
        if path[0] != '':
            with open(path[0], 'w') as f:
                f.write(self.results_textbox.toPlainText())
            self.popupmsg('Results were saved successfully!', 'notice')

    def clear_history(self) -> None:
        self.results_textbox.clear()
        self.fit_number = 0

    def get_fit_function_file_name(self) -> str:
        return self.lineEdit_pathmodel.text().split('/')[-1]

    def disable_dx(self, index) -> None:
        self.spinBox_dxcol.setEnabled(not index)

    def get_method(self) -> str:
        """
        return a string which the Fit class recognizes.
        """
        method = self.comboBox_method.currentText()
        if method == 'ODR':
            return 'odr'
        return 'ls'

    def get_data_file_ext(self) -> str:
        """
        return the file extenstion including the '.'
        """
        return '.' + self.lineEdit_pathdata.text().split('/')[-1].split('.')[-1]

    def browsefilesdata(self) -> None:
        fname = QFileDialog.getOpenFileName(self,
                                            'Open File',
                                            self.default_data_path,
                                            'CSV Files (*.csv);;Excel Files (*.xlsx *.xls *.xlsm *.xlsb *.odf *.ods *.odt)'
                                            )

        self.data_fname = fname[0]

        self.sheets = 0  # the default sheet number that pandas take is 0
        if self.data_fname.endswith(('.xlsx', '.xlsm')):
            from openpyxl import load_workbook

            wb = load_workbook(self.data_fname, read_only=True)
            sheets = wb.sheetnames

            self.comboBox_sheets.clear()  # Incase a xlsx file is loaded one after another
            self.comboBox_sheets.addItems(sheets)

            self.label_sheets.show()
            self.comboBox_sheets.show()
        else:  # Incase the user loads a non xlsx file after a xlsx file has been loaded
            self.comboBox_sheets.clear()
            self.label_sheets.hide()
            self.comboBox_sheets.hide()

        self.lineEdit_pathdata.setText(self.data_fname)

    def browsefilesmodel(self) -> None:

        if self.comboBox_method.currentText() == 'ODR':
            fname = QFileDialog.getOpenFileName(self,
                                                'Open File',
                                                self.default_odr_path,
                                                'Python Script (*.py)'
                                                )
        else:
            fname = QFileDialog.getOpenFileName(self,
                                                'Open File',
                                                self.default_ls_path,
                                                'Python Script (*.py)'
                                                )
        self.lineEdit_pathmodel.setText(fname[0])

    def set_default_path(self, field: str) -> None:
        directory = str(QFileDialog.getExistingDirectory(self, 'Choose Directory'))

        if exists(self.config_path):
            with open(self.config_path, 'r') as f:
                def_paths = json.load(f)

            def_paths[field] = directory
            with open(self.config_path, 'w') as f:
                json.dump(def_paths, f, indent=4, separators=(", ", ": "), sort_keys=True)

        else:
            with open(self.config_path, 'w') as fp:
                json.dump({field: directory}, fp, indent=4, separators=(", ", ": "), sort_keys=True)

        self.popupmsg('Path was saved successfully!', 'notice')

    def load_fit_function(self) -> None:
        module = import_source_file(self.lineEdit_pathmodel.text(), 'fit_function')
        self.fit_function = module.fit_function

    def check_identical_cols_nums(self) -> None:
        if self.method == 'ls':
            if self.xcol == self.ycol or self.xcol == self.dycol or self.dycol == self.ycol:
                raise Exception('columns must be different for:\n x, y, dy')

        if self.xcol == self.ycol \
                or self.xcol == self.dycol \
                or self.dycol == self.ycol \
                or self.xcol == self.dxcol \
                or self.dxcol == self.ycol \
                or self.dxcol == self.dycol:
            raise Exception('columns must be different for:\nx, y, dx, dy')

    def show_current_input(self) -> None:
        input = f'Data Path: {self.lineEdit_pathdata.text()} -> DType: {type(self.lineEdit_pathdata.text())}\n' \
                f'Model Path: {self.lineEdit_pathmodel.text()} -> DType: {type(self.lineEdit_pathmodel.text())}\n' \
                f'Method: {self.comboBox_method.currentText()} -> DType: {type(self.comboBox_method.currentText())}\n' \
                f'Headers: {self.checkBox_headers.isChecked()} -> DType: {type(self.checkBox_headers.isChecked())}\n' \
                f'X Column: {self.spinBox_xcol.value()} -> DType: {type(self.spinBox_xcol.value())}\n' \
                f'Y Column: {self.spinBox_ycol.value()} -> DType: {type(self.spinBox_ycol.value())}\n' \
                f'dX Column: {self.spinBox_dxcol.value()} -> DType: {type(self.spinBox_dxcol.value())}\n' \
                f'dY Column: {self.spinBox_dxcol.value()} -> DType: {type(self.spinBox_dxcol.value())}\n' \
                f'Delete Points?: {self.checkBox_delpoints.isChecked()} -> DType: {type(self.checkBox_delpoints.isChecked())}\n' \
                f'Points to Remove: {self.lineEdit_listpoints.text()} -> DType: {type(self.lineEdit_listpoints.text())}\n' \
                f'Initial Parameters: {self.lineEdit_params.text()} -> DType: {type(self.lineEdit_params.text())}\n' \
                f'Fit Title: {self.lineEdit_fittitle.text()} -> DType: {type(self.lineEdit_fittitle.text())}\n' \
                f'Fit X: {self.lineEdit_fitxlabel.text()} -> DType: {type(self.lineEdit_fitxlabel.text())}\n' \
                f'Fit Y: {self.lineEdit_fitylabel.text()} -> DType: {type(self.lineEdit_fitylabel.text())}\n' \
                f'Res Y: {self.lineEdit_resylabel.text()} -> DType: {type(self.lineEdit_resylabel.text())}'
        msg = QMessageBox()
        msg.setText(input)
        msg.exec_()

    def check_empty_fields(self) -> None:

        flag = False
        string = ''
        if self.lineEdit_pathdata.text() == '':
            string += 'Data file path is missing\n'
            flag = True
        if self.lineEdit_pathmodel.text() == '':
            string += 'Model file path is missing\n'
            flag = True
        if self.lineEdit_listpoints.text() == '' and self.checkBox_delpoints.isChecked():
            string += 'Points to remove are missing\n'
            flag = True
        if self.lineEdit_params.text() == '':
            string += 'Initial parameters are missing\n'
            flag = True
        if self.lineEdit_fittitle.text() == '':
            string += 'Fit title is missing\n'
            flag = True
        if self.lineEdit_fitxlabel.text() == '':
            string += 'Fit X label is missing\n'
            flag = True
        if self.lineEdit_fitylabel.text() == '':
            string += 'Fit Y label is missing\n'
            flag = True
        if self.lineEdit_resylabel.text() == '':
            string += 'Residuals Y label is missing\n'
            flag = True

        if flag:
            raise ValueError(string)

    def popupmsg(self, text: str, type: str) -> None:
        """
        supported types:
        * 'notice'
        * 'help'
        * 'error'

        """
        msg = QMessageBox()

        if type == 'notice':
            msg.setWindowTitle('Notice')
            msg.setIcon(QMessageBox.Information)
            msg.setInformativeText(text)
        elif type == 'help':
            msg.setWindowTitle('Help')
            msg.setIcon(QMessageBox.Question)
            msg.setInformativeText(text)
        elif type == 'error':
            msg.setWindowTitle('Error')
            msg.setIcon(QMessageBox.Critical)
            msg.setInformativeText(text)

        msg.exec_()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = FitGUI()

    sys.exit(app.exec_())
